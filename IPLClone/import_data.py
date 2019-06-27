import click
import openpyxl
import MySQLdb
import os
import csv

import django
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'IPLClone.settings')
django.setup()
from IPLproject.models import *


@click.group()
@click.pass_context
def cli(ctx):
    pass

def connectDB(db_name):
    con = MySQLdb.connect("localhost", "root", "04111998",db_name)
    return con
    pass


def dropdb(ctx,db_name):
    con=connectDB(db_name)
    cur = con.cursor()
    cur.execute("DROP DATABASE IF EXISTS onlineclass")
    con.commit()
    pass

@cli.command()
@click.argument('source_excel1', nargs=1)
@click.argument('source_excel2', nargs=1)
@click.pass_context
def importdata(ctx,source_excel1, source_excel2):
    print(source_excel2)

    # wb1=openpyxl.load_workbook(source_excel1)
    #
    # sheet = wb1.get_active_sheet()
    # rows = sheet.max_row
    # columns = sheet.max_column
    # for r in range(2, rows + 1):
    #     val = []
    #     for c in range(1, columns + 1):
    #         e = sheet.cell(row=r, column=c)
    #         val.append(e.value)
    #     valt = tuple(val)
    #     match = Match(
    #         id = valt[0],
    #         season=valt[1],
    #         city = valt[2],
    #         date = valt[3],
    #         team1 = valt[4],
    #         team2 = valt[5],
    #         toss_winner = valt[6],
    #         toss_decision = valt[7],
    #         result = valt[8],
    #         dl_applied = valt[9],
    #         winner = valt[10],
    #         win_by_runs = valt[11],
    #         win_by_wickets = valt[12],
    #         player_of_match = valt[13],
    #         venue = valt[14],
    #         umpire1 = valt[15],
    #         umpire2 = valt[16],
    #         umpire3 = valt[17]
    #     )
    #     match.save()

    # wb2 = openpyxl.load_workbook(source_excel2)
    #
    # sheet2 = wb2.get_active_sheet()
    # rows2=sheet2.max_row
    # columns2=sheet2.max_column
    # print(rows2, columns2)
    # for r in range(2,rows2+1):
    #     val=[]
    #     for c in range(1,columns2+1):
    #         e=sheet2.cell(row=r,column=c)
    #         val.append(e.value)
    #     valt=tuple(val)
    #     deliveries = Deliveries(
    #         match = Match.objects.get(id=valt[0]),
    #         inning = valt[1],
    #         batting_team = valt[2],
    #         bowling_team = valt[3],
    #         over = valt[4],
    #         ball = valt[5],
    #         batsman = valt[6],
    #         non_striker = valt[7],
    #         bowler = valt[8],
    #         is_super_over = valt[9],
    #         wide_runs = valt[10],
    #         bye_runs = valt[11],
    #         legbye_runs = valt[12],
    #         noball_runs = valt[13],
    #         penalty_runs = valt[14],
    #         batsman_runs = valt[15],
    #         extra_runs = valt[16],
    #         total_runs = valt[17],
    #         player_dismissed = valt[18],
    #         dismissal_kind = valt[19],
    #         fielder = valt[20]
    #     )
    #     deliveries.save()
    with open('deliveries.csv', 'rt') as f:
        data = csv.DictReader(f)
        for row in data:
            match = Deliveries(**row)
            match.save()


if __name__ == '__main__':
    cli(obj={})